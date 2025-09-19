import streamlit as st
import sqlite3
import pandas as pd
import os
from datetime import datetime, timedelta
import re
from io import BytesIO

# --- 追加: Excel出力に必要なライブラリ ---
import shutil
import openpyxl

# データベースパス
DB_PATH = os.path.join(os.path.dirname(__file__), "skillsheet_data.db")

st.title("📝 スキルシート更新")

# ページ内ナビゲーション
nav_cols = st.columns([1, 1, 8])
with nav_cols[0]:
    if st.button("🏠 ホームへ戻る", key="go_home_from_update"):
        st.session_state.current_page = "🏠 ホーム"
        st.rerun()

st.markdown("---")

# 編集モードの選択
edit_mode = st.radio(
    "編集モードを選択してください:",
    ["基本情報の編集", "案件情報の編集", "新しい案件情報の追加"],
    key="edit_mode"
)

st.markdown("---")

# データベースからユーザー一覧を取得
def get_user_list():
    """ユーザー一覧を取得"""
    conn = sqlite3.connect(DB_PATH)
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_info'")
        if cursor.fetchone():
            query = """
                SELECT id, name, name_kana, created_at
                FROM user_info
                ORDER BY created_at DESC
            """
        else:
            query = """
                SELECT id, name, name_kana, created_at
                FROM basic_info
                ORDER BY created_at DESC
            """
        df = pd.read_sql_query(query, conn)
        return df
    except Exception as e:
        st.error(f"ユーザー一覧の取得に失敗しました: {str(e)}")
        return pd.DataFrame()
    finally:
        conn.close()

# ユーザー選択
st.subheader("更新するユーザーを選択してください")
user_df = get_user_list()

if user_df.empty:
    st.warning("データベースにユーザー情報がありません。")
else:
    # ユーザー選択用のセレクトボックス
    user_options = []
    for _, row in user_df.iterrows():
        display_name = f"{row['name']} ({row['name_kana']}) - {row['created_at']}"
        user_options.append((display_name, row['id']))
    
    selected_user = st.selectbox(
        "ユーザーを選択:",
        options=[opt[1] for opt in user_options],
        format_func=lambda x: next(opt[0] for opt in user_options if opt[1] == x),
        key="selected_user"
    )
    
    if selected_user:
        # ユーティリティ: スキル名の正規化（記号除去・空白トリム）
        def normalize_skill_name(name):
            if name is None:
                return ""
            s = str(name).strip()
            # 記号除去: [ ] ' " など
            s = re.sub(r"^[\[\]'\"]+|[\[\]'\"]+$", "", s)
            s = s.strip()
            return s

        # 年月→「〇年〇ヶ月」表記
        def float_to_year_month(val):
            try:
                total_months = int(round(float(val) * 12))
                years = total_months // 12
                months = total_months % 12
                if years > 0 and months > 0:
                    return f"{years}年{months}ヶ月"
                elif years > 0:
                    return f"{years}年"
                elif months > 0:
                    return f"{months}ヶ月"
                else:
                    return "0ヶ月"
            except Exception:
                return ""

        # 「〇年〇ヶ月」→ float年
        def year_month_to_float(years, months):
            try:
                return float(years) + float(months) / 12.0
            except Exception:
                return 0.0

        # float年→(年, 月)
        def float_to_years_months(val):
            try:
                total_months = int(round(float(val) * 12))
                years = total_months // 12
                months = total_months % 12
                return years, months
            except Exception:
                return 0, 0

        # 選択されたユーザーの詳細情報を取得
        def get_user_details(user_id):
            conn = sqlite3.connect(DB_PATH)
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_info'")
                if cursor.fetchone():
                    # 新しいスキーマ
                    user_query = "SELECT * FROM user_info WHERE id = ?"
                    skills_query = "SELECT * FROM skills WHERE user_info_id = ?"
                    projects_query = "SELECT * FROM projects WHERE user_info_id = ?"
                else:
                    # 古いスキーマ
                    user_query = "SELECT * FROM basic_info WHERE id = ?"
                    skills_query = None
                    projects_query = "SELECT * FROM projects WHERE basic_info_id = ?"
                user_data = pd.read_sql_query(user_query, conn, params=(user_id,))
                skills_data = pd.DataFrame()
                if skills_query:
                    skills_data = pd.read_sql_query(skills_query, conn, params=(user_id,))
                projects_data = pd.read_sql_query(projects_query, conn, params=(user_id,))
                return user_data, skills_data, projects_data
            except Exception as e:
                st.error(f"ユーザー詳細の取得に失敗しました: {str(e)}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            finally:
                conn.close()
        
        user_data, skills_data, projects_data = get_user_details(selected_user)

        # --- Excel出力用関数 ---
        def export_user_to_excel(user_data, skills_data, projects_data):
            # 必要なライブラリは冒頭でimport済み
            def safe_write_cell(ws, cell_address, value):
                try:
                    cell = ws[cell_address]
                    if hasattr(cell, 'value') and hasattr(cell, 'coordinate'):
                        cell.value = value
                    else:
                        if hasattr(cell, 'coordinate'):
                            master_cell = ws[cell.coordinate]
                            master_cell.value = value
                except Exception:
                    pass

            def clean_project_value(val):
                if val is None:
                    return ""
                if isinstance(val, str):
                    s = val.strip()
                    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
                        s = s[1:-1]
                    if s.startswith("[") and s.endswith("]"):
                        try:
                            import ast
                            parsed = ast.literal_eval(s)
                            if isinstance(parsed, list):
                                return ", ".join(str(x).strip('"').strip("'") for x in parsed if str(x).strip() and x not in ["", None, "''", '""', "[]"])
                            else:
                                return str(parsed)
                        except Exception:
                            return s[1:-1]
                    return s
                if isinstance(val, list):
                    return ", ".join(str(x).strip('"').strip("'") for x in val if str(x).strip() and x not in ["", None, "''", '""', "[]"])
                return str(val)

            def split_env_items(val):
                if val is None:
                    return []
                if isinstance(val, dict):
                    return []
                if isinstance(val, list):
                    return [str(x).strip() for x in val if str(x).strip() and x not in ["", None, "''", '""', "[]"]]
                s = str(val).strip()
                if s.startswith("[") and s.endswith("]"):
                    try:
                        import ast
                        parsed = ast.literal_eval(s)
                        if isinstance(parsed, list):
                            return [str(x).strip('"').strip("'").strip() for x in parsed if str(x).strip() and x not in ["", None, "''", '""', "[]"]]
                    except Exception:
                        pass
                items = []
                for part in s.replace('\r\n', '\n').replace('\r', '\n').replace(',', '\n').split('\n'):
                    part = part.strip()
                    if part and part not in ['""', "''", "[]"]:
                        if part.strip() not in ["", "''", '""', "[]"]:
                            items.append(part)
                return items

            def split_qualifications(qual_str):
                if not qual_str:
                    return [], []
                items = []
                for part in str(qual_str).replace('\r\n', '\n').replace('\r', '\n').replace(',', '\n').split('\n'):
                    part = part.strip()
                    if part and part not in ['""', "''", "[]"]:
                        items.append(part)
                first6 = items[:6]
                rest = items[6:]
                return first6, rest

            # --- テンプレートファイルのパスを Template フォルダ配下に変更 ---
            template_path = os.path.join(os.path.dirname(__file__), "Template", "SkillSheetTemplate.xlsx")
            if not os.path.exists(template_path):
                st.error("テンプレートファイル（Template/SkillSheetTemplate.xlsx）が見つかりません。")
                st.info(
                    "テンプレートファイル（Template/SkillSheetTemplate.xlsx）がこのアプリのTemplateフォルダに存在しません。"
                    "テンプレートが必要な場合は、管理者または開発者に連絡してテンプレートファイルを入手し、"
                    "このアプリのTemplateフォルダに配置してください。"
                )
                return None, None

            # 生成ファイルの保存先を明示的に指定し、Streamlitのダウンロードボタンと同時に物理ファイルも残す
            output_dir = os.path.join(os.path.dirname(__file__), "generated_excels")
            os.makedirs(output_dir, exist_ok=True)
            output_filename = f"SkillSheetOutput_{user_data.iloc[0]['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(output_dir, output_filename)

            try:
                shutil.copy(template_path, output_path)
                wb = openpyxl.load_workbook(output_path)
                ws = wb.active
                try:
                    ws["AG39"] = ""
                except Exception:
                    pass

                if not user_data.empty:
                    user = user_data.iloc[0]
                    safe_write_cell(ws, "E5", user.get('name', ''))
                    safe_write_cell(ws, "E4", user.get('name_kana', ''))
                    safe_write_cell(ws, "T5", user.get('transportation', ''))
                    safe_write_cell(ws, "AB5", user.get('nearest_station', ''))
                    safe_write_cell(ws, "AJ5", user.get('access_method', ''))
                    safe_write_cell(ws, "AM5", user.get('access_time', ''))
                    safe_write_cell(ws, "Q5", user.get('gender', ''))
                    safe_write_cell(ws, "AV1", user.get('birth_date', ''))
                    safe_write_cell(ws, "B8", user.get('final_education', ''))
                    safe_write_cell(ws, "AE8", user.get('graduation_date', ''))
                    safe_write_cell(ws, "B27", user.get('self_pr', ''))
                    qual_items, qual_rest = [], []
                    if 'qualifications' in user:
                        qual_items, qual_rest = split_qualifications(user.get('qualifications', ''))
                    qual_cells = ["B10", "B11", "B12", "V10", "V11", "V12"]
                    for i, cell in enumerate(qual_cells):
                        val = qual_items[i] if i < len(qual_items) else ""
                        safe_write_cell(ws, cell, val)
                    if qual_rest:
                        extra_cells = []
                        for n in range(13, 30):
                            extra_cells.append(f"B{n}")
                            extra_cells.append(f"V{n}")
                        for i, val in enumerate(qual_rest):
                            if i < len(extra_cells):
                                safe_write_cell(ws, extra_cells[i], val)
                    safe_write_cell(ws, "B29", "")

                skill_types = ['language', 'tool', 'db', 'machine']
                skill_cell_map = {
                    'language': {'name': 'B', 'years': 'H', 'row_start': 15},
                    'tool': {'name': 'L', 'years': 'R', 'row_start': 15},
                    'db': {'name': 'V', 'years': 'AB', 'row_start': 15},
                    'machine': {'name': 'AF', 'years': 'AL', 'row_start': 15},
                }
                for skill_type in skill_types:
                    cell_info = skill_cell_map[skill_type]
                    if not skills_data.empty:
                        type_skills = skills_data[skills_data['skill_type'] == skill_type]
                        for i in range(10):
                            if i < len(type_skills):
                                skill = type_skills.iloc[i]
                                skill_name = normalize_skill_name(skill['skill_name'])
                                years_val = skill['experience_years']
                                years_str = float_to_year_month(years_val)
                                safe_write_cell(ws, f"{cell_info['name']}{cell_info['row_start']+i}", skill_name)
                                safe_write_cell(ws, f"{cell_info['years']}{cell_info['row_start']+i}", years_str)
                            else:
                                safe_write_cell(ws, f"{cell_info['name']}{cell_info['row_start']+i}", "")
                                safe_write_cell(ws, f"{cell_info['years']}{cell_info['row_start']+i}", "")
                    else:
                        for i in range(10):
                            safe_write_cell(ws, f"{cell_info['name']}{cell_info['row_start']+i}", "")
                            safe_write_cell(ws, f"{cell_info['years']}{cell_info['row_start']+i}", "")

                project_base_rows = [37, 50, 63]
                project_cell_map = {
                    "period_start": lambda base: f"D{base}",
                    "period_end": lambda base: f"E{base+1}",
                    "system_name": lambda base: f"J{base+3}",
                    "role": lambda base: f"T{base+2}",
                    "industry": lambda base: f"L{base+2}",
                    "headcount": lambda base: f"Z{base+2}",
                    "env_langs": lambda base, i: f"AD{base+2+i}",
                    "env_dbs": lambda base, i: f"AJ{base+2+i}",
                    "env_tools": lambda base, i: f"AG{base+2+i}",
                    "env_oss": lambda base, i: f"AM{base+2+i}",
                    "work_content": lambda base: f"B{base+3}",
                    "phases": lambda base: f"AP{base+1}",
                }
                max_projects = 3

                projects_list = []
                if not projects_data.empty:
                    if 'period_start' in projects_data.columns:
                        try:
                            projects_data_sorted = projects_data.copy()
                            projects_data_sorted['period_start_sort'] = pd.to_datetime(projects_data_sorted['period_start'], errors='coerce')
                            projects_data_sorted = projects_data_sorted.sort_values('period_start_sort', ascending=False)
                            projects_list = [row for _, row in projects_data_sorted.iterrows()]
                        except Exception:
                            projects_list = [row for _, row in projects_data.sort_values('id', ascending=False).iterrows()]
                    else:
                        projects_list = [row for _, row in projects_data.sort_values('id', ascending=False).iterrows()]
                for proj_idx in range(max_projects):
                    base_row = project_base_rows[proj_idx]
                    if proj_idx < len(projects_list):
                        project = projects_list[proj_idx]
                        safe_write_cell(ws, project_cell_map["period_start"](base_row), clean_project_value(project.get('period_start', '')))
                        safe_write_cell(ws, project_cell_map["period_end"](base_row), clean_project_value(project.get('period_end', '')))
                        safe_write_cell(ws, project_cell_map["system_name"](base_row), clean_project_value(project.get('system_name', '')))
                        safe_write_cell(ws, project_cell_map["role"](base_row), clean_project_value(project.get('role', '')))
                        safe_write_cell(ws, project_cell_map["industry"](base_row), clean_project_value(project.get('industry', '')))
                        safe_write_cell(ws, project_cell_map["headcount"](base_row), clean_project_value(project.get('headcount', '')))
                        safe_write_cell(ws, project_cell_map["work_content"](base_row), clean_project_value(project.get('work_content', '')))
                        phases_val = project.get('phases', '')
                        phase_labels = [
                            "環境構築", "要件", "基本", "詳細", "製造", "単体", "結合", "総合", "保守運用", "他"
                        ]
                        phase_columns = ["J", "L", "N", "P", "R", "T", "V", "X", "Z", "AB"]

                        def is_phase_checked(val, label):
                            try:
                                if isinstance(val, dict):
                                    return bool(val.get(label, False))
                                if isinstance(val, list):
                                    return label in val
                                s = str(val)
                                if s.startswith("[") and s.endswith("]"):
                                    import ast
                                    try:
                                        parsed = ast.literal_eval(s)
                                        if isinstance(parsed, list):
                                            return label in [str(x) for x in parsed]
                                    except Exception:
                                        pass
                                for part in s.replace('\n', ',').split(','):
                                    if part.strip() == label:
                                        return True
                            except Exception:
                                return False
                            return False

                        target_row = base_row + 1
                        for col_letter, label in zip(phase_columns, phase_labels):
                            ws[f"{col_letter}{target_row}"] = "●" if is_phase_checked(phases_val, label) else ""

                        def safe_split_env_items(val):
                            if isinstance(val, dict):
                                return []
                            try:
                                import math
                                if isinstance(val, float) and math.isnan(val):
                                    return []
                            except Exception:
                                pass
                            items = split_env_items(val)
                            return [x for x in items if x and x.strip() not in {"nan", "none", "", "''", '""', "[]"}]
                        
                        env_langs_list = safe_split_env_items(project.get('env_langs', ''))
                        env_dbs_list = safe_split_env_items(project.get('env_dbs', ''))
                        env_tools_list = safe_split_env_items(project.get('env_tools', ''))
                        env_oss_list = safe_split_env_items(project.get('env_oss', ''))

                        for i in range(3):
                            val = env_langs_list[i] if i < len(env_langs_list) else ""
                            safe_write_cell(ws, project_cell_map["env_langs"](base_row, i), val)
                        for i in range(3):
                            val = env_dbs_list[i] if i < len(env_dbs_list) else ""
                            safe_write_cell(ws, project_cell_map["env_dbs"](base_row, i), val)
                        for i in range(3):
                            val = env_tools_list[i] if i < len(env_tools_list) else ""
                            safe_write_cell(ws, project_cell_map["env_tools"](base_row, i), val)
                        for i in range(3):
                            val = env_oss_list[i] if i < len(env_oss_list) else ""
                            safe_write_cell(ws, project_cell_map["env_oss"](base_row, i), val)
                    else:
                        safe_write_cell(ws, project_cell_map["period_start"](base_row), "")
                        safe_write_cell(ws, project_cell_map["period_end"](base_row), "")
                        safe_write_cell(ws, project_cell_map["system_name"](base_row), "")
                        safe_write_cell(ws, project_cell_map["role"](base_row), "")
                        safe_write_cell(ws, project_cell_map["industry"](base_row), "")
                        safe_write_cell(ws, project_cell_map["headcount"](base_row), "")
                        safe_write_cell(ws, project_cell_map["work_content"](base_row), "")
                        safe_write_cell(ws, project_cell_map["phases"](base_row), "")
                        for i in range(3):
                            safe_write_cell(ws, project_cell_map["env_langs"](base_row, i), "")
                            safe_write_cell(ws, project_cell_map["env_dbs"](base_row, i), "")
                            safe_write_cell(ws, project_cell_map["env_tools"](base_row, i), "")
                            safe_write_cell(ws, project_cell_map["env_oss"](base_row, i), "")

                wb.save(output_path)
                # 物理ファイルのパスをStreamlitで明示表示
                st.info(f"生成されたExcelファイル: `{output_path}`")
                with open(output_path, 'rb') as f:
                    output = BytesIO(f.read())
                # ファイルは削除しない（物理的に残す）
                return output, output_path

            except Exception as e:
                st.error(f"Excel出力中にエラーが発生しました: {str(e)}")
                return None, None

        # Excel出力ボタン
        st.markdown("### データ出力")
        excel_btn_col, _ = st.columns([1, 9])
        with excel_btn_col:
            if st.button("このユーザーの情報をExcel出力", key="export_excel"):
                excel_bytes, output_path = export_user_to_excel(user_data, skills_data, projects_data)
                if excel_bytes:
                    st.download_button(
                        label="ダウンロード（Excel）",
                        data=excel_bytes,
                        file_name=f"{user_data.iloc[0]['name']}SkillSheet.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success(f"Excelファイルが生成されました: `{output_path}`")
                else:
                    st.error("Excelファイルの生成に失敗しました。")

        # --- 以下は元のまま ---
        if not user_data.empty:
            # 基本情報の表示と編集
            if edit_mode == "基本情報の編集":
                st.subheader("基本情報の編集")
                with st.form("edit_basic_info_form"):
                    col1, col2 = st.columns(2)
                    with col1:
                        name = st.text_input("氏名", value=user_data.iloc[0]['name'], key="edit_name")
                        name_kana = st.text_input("カナ", value=user_data.iloc[0]['name_kana'], key="edit_name_kana")
                        gender = st.selectbox("性別", ["男", "女"], index=0 if user_data.iloc[0]['gender'] == "男" else 1, key="edit_gender")
                        birth_date_value = None
                        if user_data.iloc[0]['birth_date']:
                            try:
                                birth_date_str = str(user_data.iloc[0]['birth_date'])
                                if '/' in birth_date_str:
                                    birth_date_value = datetime.strptime(birth_date_str, '%Y/%m/%d').date()
                                else:
                                    birth_date_value = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                            except:
                                birth_date_value = None
                        birth_date = st.date_input("生年月日", value=birth_date_value, key="edit_birth_date")
                        final_education = st.text_input("最終学歴", value=user_data.iloc[0]['final_education'], key="edit_final_education")
                    with col2:
                        transportation = st.text_input("電車", value=user_data.iloc[0]['transportation'], key="edit_transportation")
                        nearest_station = st.text_input("最寄り駅", value=user_data.iloc[0]['nearest_station'], key="edit_nearest_station")
                        access_method = st.selectbox("駅からの交通手段", ["徒歩", "自転車", "バス", "車"], 
                                                   index=["徒歩", "自転車", "バス", "車"].index(user_data.iloc[0]['access_method']) if user_data.iloc[0]['access_method'] in ["徒歩", "自転車", "バス", "車"] else 0, 
                                                   key="edit_access_method")
                        access_time = st.text_input("所要時間(分)", value=user_data.iloc[0]['access_time'], key="edit_access_time")
                        graduation_date = st.text_input("卒業年月", value=user_data.iloc[0]['graduation_date'], key="edit_graduation_date")
                    self_pr = st.text_area("自己PR", value=user_data.iloc[0]['self_pr'], key="edit_self_pr")
                    qualifications = st.text_area("資格", value=user_data.iloc[0]['qualifications'], key="edit_qualifications")
                    submitted = st.form_submit_button("基本情報を更新", type="primary")
                    if submitted:
                        def update_basic_info(user_id, updated_data):
                            conn = sqlite3.connect(DB_PATH)
                            try:
                                cursor = conn.cursor()
                                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_info'")
                                use_new_schema = cursor.fetchone() is not None
                                if use_new_schema:
                                    cursor.execute("""
                                        UPDATE user_info SET 
                                        name = ?, name_kana = ?, transportation = ?, nearest_station = ?,
                                        access_method = ?, access_time = ?, gender = ?, birth_date = ?,
                                        final_education = ?, graduation_date = ?, self_pr = ?, qualifications = ?
                                        WHERE id = ?
                                    """, (
                                        updated_data['name'], updated_data['name_kana'], updated_data['transportation'],
                                        updated_data['nearest_station'], updated_data['access_method'], updated_data['access_time'],
                                        updated_data['gender'], updated_data['birth_date'], updated_data['final_education'],
                                        updated_data['graduation_date'], updated_data['self_pr'], updated_data['qualifications'],
                                        user_id
                                    ))
                                else:
                                    cursor.execute("""
                                        UPDATE basic_info SET 
                                        name = ?, name_kana = ?, transportation = ?, nearest_station = ?,
                                        access_method = ?, access_time = ?, gender = ?, birth_date = ?,
                                        final_education = ?, graduation_date = ?, self_pr = ?
                                        WHERE id = ?
                                    """, (
                                        updated_data['name'], updated_data['name_kana'], updated_data['transportation'],
                                        updated_data['nearest_station'], updated_data['access_method'], updated_data['access_time'],
                                        updated_data['gender'], updated_data['birth_date'], updated_data['final_education'],
                                        updated_data['graduation_date'], updated_data['self_pr'], user_id
                                    ))
                                conn.commit()
                                return True
                            except Exception as e:
                                conn.rollback()
                                st.error(f"基本情報の更新に失敗しました: {str(e)}")
                                return False
                            finally:
                                conn.close()
                        updated_data = {
                            'name': name,
                            'name_kana': name_kana,
                            'transportation': transportation,
                            'nearest_station': nearest_station,
                            'access_method': access_method,
                            'access_time': access_time,
                            'gender': gender,
                            'birth_date': birth_date.strftime('%Y-%m-%d') if birth_date else None,
                            'final_education': final_education,
                            'graduation_date': graduation_date,
                            'self_pr': self_pr,
                            'qualifications': qualifications
                        }
                        if update_basic_info(selected_user, updated_data):
                            st.success("基本情報を更新しました！")
                            st.rerun()
                        else:
                            st.error("基本情報の更新に失敗しました。")
            else:
                st.subheader("基本情報")
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**氏名:** {user_data.iloc[0]['name']}")
                    st.write(f"**カナ:** {user_data.iloc[0]['name_kana']}")
                    st.write(f"**性別:** {user_data.iloc[0]['gender']}")
                with col2:
                    st.write(f"**最寄り駅:** {user_data.iloc[0]['nearest_station']}")
                    st.write(f"**交通手段:** {user_data.iloc[0]['access_method']}")
                    st.write(f"**所要時間:** {user_data.iloc[0]['access_time']}分")
            st.markdown("---")
            
            # 案件情報の表示・編集
            if edit_mode == "案件情報の編集":
                st.subheader("案件情報の編集")
                
                # --- 案件情報の出力順を「新しい順」にする ---
                projects_list = []
                if not projects_data.empty:
                    if 'period_start' in projects_data.columns:
                        try:
                            projects_data_sorted = projects_data.copy()
                            projects_data_sorted['period_start_sort'] = pd.to_datetime(projects_data_sorted['period_start'], errors='coerce')
                            projects_data_sorted = projects_data_sorted.sort_values('period_start_sort', ascending=False)
                            projects_list = [row for _, row in projects_data_sorted.iterrows()]
                        except Exception:
                            projects_list = [row for _, row in projects_data.sort_values('id', ascending=False).iterrows()]
                    else:
                        projects_list = [row for _, row in projects_data.sort_values('id', ascending=False).iterrows()]

                # 編集フォームの値を一時保存するためのセッションステート
                if "edit_projects_buffer" not in st.session_state or st.session_state.get("edit_projects_buffer_user") != selected_user:
                    st.session_state["edit_projects_buffer"] = []
                    for project in projects_list:
                        st.session_state["edit_projects_buffer"].append(dict(project))
                    st.session_state["edit_projects_buffer_user"] = selected_user

                # 編集用バッファ
                edit_projects_buffer = st.session_state["edit_projects_buffer"]

                st.write("**現在の案件情報:**")
                expander_states = []
                for idx, project in enumerate(edit_projects_buffer):
                    with st.expander(f"案件 {idx + 1}: {project.get('system_name', 'N/A')} - 編集", expanded=True):
                        # --- 編集フォーム ---
                        col1, col2 = st.columns(2)
                        with col1:
                            def parse_ym(val):
                                if not val:
                                    return ""
                                if isinstance(val, str):
                                    try:
                                        if "/" in val:
                                            return datetime.strptime(val, "%Y/%m").strftime("%Y/%m")
                                        elif "-" in val:
                                            return datetime.strptime(val, "%Y-%m").strftime("%Y/%m")
                                        elif len(val) == 6 and val.isdigit():
                                            return f"{val[:4]}/{val[4:]}"
                                    except:
                                        return val
                                elif isinstance(val, datetime):
                                    return val.strftime("%Y/%m")
                                return str(val)
                            period_start_ym = parse_ym(project.get('period_start', ''))
                            period_end_ym = parse_ym(project.get('period_end', ''))
                            period_start = st.text_input("開始年月 (yyyy/MM)", value=period_start_ym, key=f"edit_proj_{idx}_period_start")
                            period_end = st.text_input("終了年月 (yyyy/MM)", value=period_end_ym, key=f"edit_proj_{idx}_period_end")
                            system_name = st.text_input("システム名・作業内容", value=project.get('system_name', ''), key=f"edit_proj_{idx}_system_name")
                            role_options = [
                                "CNSL(コンサルタント)", "PMO(プロジェクトマネジメントオフィス)", 
                                "PM(プロジェクトマネージャー)", "PL(プロジェクトリーダー)",
                                "SL(サブリーダー)", "SE(システムエンジニア)", "PG(プログラマ)", "M（メンバー）"
                            ]
                            current_role = project.get('role', '')
                            role_index = role_options.index(current_role) if current_role in role_options else 0
                            role = st.selectbox("役割", role_options, index=role_index, key=f"edit_proj_{idx}_role")
                        with col2:
                            industry_options = [
                                "農林業", "鉱業", "建設業", "製造業", "情報通信", "運輸業", 
                                "卸売・小売", "金融・保険", "不動産業", "飲食・宿泊所", 
                                "医療・福祉", "教育・学習", "複合サービス事業", "サービス業", "公務", "その他"
                            ]
                            current_industry = project.get('industry', '')
                            industry_index = industry_options.index(current_industry) if current_industry in industry_options else 0
                            industry = st.selectbox("業種", industry_options, index=industry_index, key=f"edit_proj_{idx}_industry")
                            headcount = st.text_input("人数", value=project.get('headcount', ''), key=f"edit_proj_{idx}_headcount")
                        st.write("**工程:**")
                        phase_cols = st.columns(5)
                        current_phases = []
                        phases_raw = project.get('phases', '')
                        if phases_raw:
                            try:
                                if isinstance(phases_raw, list):
                                    current_phases = phases_raw
                                elif isinstance(phases_raw, str):
                                    if phases_raw.startswith('[') and phases_raw.endswith(']'):
                                        import ast
                                        try:
                                            parsed = ast.literal_eval(phases_raw)
                                            if isinstance(parsed, list):
                                                current_phases = parsed
                                            else:
                                                current_phases = []
                                        except Exception:
                                            current_phases = []
                                    else:
                                        current_phases = [phases_raw] if phases_raw else []
                            except:
                                current_phases = []
                        phases = []
                        phase_options = ["環境構築", "要件", "基本", "詳細", "製造", "単体", "結合", "総合", "保守運用", "他"]
                        for i, phase in enumerate(phase_options):
                            with phase_cols[i % 5]:
                                checked = phase in current_phases
                                if st.checkbox(phase, value=checked, key=f"edit_proj_{idx}_phase_{i}"):
                                    phases.append(phase)
                        st.write("**環境情報:**")
                        env_cols = st.columns(4)
                        env_types = [
                            ("env_langs", "言語", "language"),
                            ("env_tools", "ツール/FW/Lib", "tool"),
                            ("env_dbs", "DB", "db"),
                            ("env_oss", "OS/マシン", "machine"),
                        ]
                        env_inputs = {}
                        env_years_dicts = {}
                        for i, (env_type, env_label, env_key) in enumerate(env_types):
                            env_val = project.get(env_type, "")
                            years_dict_key = f"edit_proj_{idx}_{env_type}_years"
                            if years_dict_key not in st.session_state:
                                env_items = [normalize_skill_name(x) for x in re.split(r"[\n,]", str(env_val)) if normalize_skill_name(x) and normalize_skill_name(x) not in ["''", '""', "[]"]]
                                # 新しい形式: 年/月入力
                                st.session_state[years_dict_key] = {x: (0, 0) for x in env_items}
                            with env_cols[i]:
                                env_text = st.text_area(env_label, value=env_val, key=f"edit_proj_{idx}_{env_type}")
                                env_items = [normalize_skill_name(x) for x in re.split(r"[\n,]", str(env_text)) if normalize_skill_name(x) and normalize_skill_name(x) not in ["''", '""', "[]"]]
                                years_dict = st.session_state.get(years_dict_key, {})
                                # 既存の値をfloat→(年,月)に変換
                                for x in env_items:
                                    if x not in years_dict:
                                        # skills_dataから既存年数を取得
                                        years, months = 0, 0
                                        if not skills_data.empty:
                                            skill_row = skills_data[(skills_data['skill_type'] == env_key) & (skills_data['skill_name'] == x)]
                                            if not skill_row.empty:
                                                y, m = float_to_years_months(skill_row.iloc[0]['experience_years'])
                                                years, months = y, m
                                        years_dict[x] = (years, months)
                                for k in list(years_dict.keys()):
                                    if k not in env_items:
                                        del years_dict[k]
                                for x in env_items:
                                    year_key = f"edit_proj_{idx}_{env_type}_years_{x}_year"
                                    month_key = f"edit_proj_{idx}_{env_type}_years_{x}_month"
                                    default_year, default_month = years_dict.get(x, (0, 0))
                                    y = st.number_input(
                                        f"{env_label}「{x}」経験年数（年）", min_value=0, max_value=50, step=1,
                                        value=st.session_state.get(year_key, default_year), key=year_key
                                    )
                                    m = st.number_input(
                                        f"{env_label}「{x}」経験年数（月）", min_value=0, max_value=11, step=1,
                                        value=st.session_state.get(month_key, default_month), key=month_key
                                    )
                                    years_dict[x] = (y, m)
                                    st.session_state[years_dict_key] = years_dict
                                    st.caption(f"→ {y}年{m}ヶ月")
                                env_inputs[env_type] = env_text
                                env_years_dicts[env_type] = years_dict

                        if st.button("案件を削除", key=f"edit_proj_{idx}_delete"):
                            def delete_project(project_id):
                                conn = sqlite3.connect(DB_PATH)
                                try:
                                    cursor = conn.cursor()
                                    cursor.execute("DELETE FROM projects WHERE id = ?", (project.get('id'),))
                                    conn.commit()
                                    return True
                                except Exception as e:
                                    conn.rollback()
                                    st.error(f"案件情報の削除に失敗しました: {str(e)}")
                                    return False
                                finally:
                                    conn.close()
                            if delete_project(project.get('id')):
                                st.success("案件情報を削除しました！")
                                del st.session_state["edit_projects_buffer"]
                                st.rerun()
                            else:
                                st.error("案件情報の削除に失敗しました。")
                        edit_projects_buffer[idx]['period_start'] = st.session_state.get(f"edit_proj_{idx}_period_start", period_start_ym)
                        edit_projects_buffer[idx]['period_end'] = st.session_state.get(f"edit_proj_{idx}_period_end", period_end_ym)
                        edit_projects_buffer[idx]['system_name'] = st.session_state.get(f"edit_proj_{idx}_system_name", project.get('system_name', ''))
                        edit_projects_buffer[idx]['role'] = st.session_state.get(f"edit_proj_{idx}_role", current_role)
                        edit_projects_buffer[idx]['industry'] = st.session_state.get(f"edit_proj_{idx}_industry", current_industry)
                        edit_projects_buffer[idx]['headcount'] = st.session_state.get(f"edit_proj_{idx}_headcount", project.get('headcount', ''))
                        edit_projects_buffer[idx]['phases'] = phases
                        for env_type, _, env_key in env_types:
                            edit_projects_buffer[idx][env_type] = st.session_state.get(f"edit_proj_{idx}_{env_type}", project.get(env_type, ''))
                            edit_projects_buffer[idx][f"{env_type}_years"] = st.session_state.get(f"edit_proj_{idx}_{env_type}_years", st.session_state.get(f"edit_proj_{idx}_{env_type}_years", {}))

                if st.button("すべての案件情報を更新", key="update_all_projects"):
                    def update_skills_with_env_items(user_id, env_type, env_items, years_dict):
                        conn = sqlite3.connect(DB_PATH)
                        try:
                            cursor = conn.cursor()
                            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_info'")
                            use_new_schema = cursor.fetchone() is not None
                            if not use_new_schema:
                                return
                            cursor.execute("SELECT skill_type, skill_name, experience_years FROM skills WHERE user_info_id = ?", (user_id,))
                            all_skills = cursor.fetchall()
                            skill_dict = {}
                            for skill_type, skill_name, exp in all_skills:
                                norm_name = normalize_skill_name(skill_name)
                                skill_dict[(skill_type, norm_name)] = exp
                            for item in env_items:
                                norm_item = normalize_skill_name(item)
                                if not norm_item or norm_item in ["''", '""', "[]"]:
                                    continue
                                y, m = years_dict.get(norm_item, (0, 0))
                                years = year_month_to_float(y, m)
                                key = (env_type, norm_item)
                                if key in skill_dict:
                                    new_years = years  # 上書き方式
                                    cursor.execute("""
                                        UPDATE skills SET experience_years = ? 
                                        WHERE user_info_id = ? AND skill_type = ? AND skill_name = ?
                                    """, (new_years, user_id, env_type, norm_item))
                                else:
                                    cursor.execute("""
                                        INSERT INTO skills (user_info_id, skill_type, skill_name, experience_years)
                                        VALUES (?, ?, ?, ?)
                                    """, (user_id, env_type, norm_item, years))
                            conn.commit()
                        except Exception as e:
                            conn.rollback()
                            st.error(f"スキル経験年数の更新に失敗しました: {str(e)}")
                        finally:
                            conn.close()
                    def update_project(project_id, updated_data):
                        conn = sqlite3.connect(DB_PATH)
                        try:
                            cursor = conn.cursor()
                            cursor.execute("""
                                UPDATE projects SET 
                                period_start = ?, period_end = ?, system_name = ?,
                                role = ?, industry = ?, work_content = ?, phases = ?,
                                headcount = ?, env_langs = ?, env_tools = ?,
                                env_dbs = ?, env_oss = ?
                                WHERE id = ?
                            """, (
                                updated_data['period_start'], updated_data['period_end'],
                                updated_data['system_name'], updated_data['role'], updated_data['industry'],
                                updated_data.get('work_content', ''), str(updated_data['phases']) if not isinstance(updated_data['phases'], dict) else '', updated_data['headcount'],
                                updated_data['env_langs'], updated_data['env_tools'],
                                updated_data['env_dbs'], updated_data['env_oss'], project_id
                            ))
                            conn.commit()
                            return True
                        except Exception as e:
                            conn.rollback()
                            st.error(f"案件情報の更新に失敗しました: {str(e)}")
                            return False
                        finally:
                            conn.close()
                    all_success = True
                    for idx, project in enumerate(edit_projects_buffer):
                        updated_data = {
                            'period_start': project.get('period_start', ''),
                            'period_end': project.get('period_end', ''),
                            'system_name': project.get('system_name', ''),
                            'role': project.get('role', ''),
                            'industry': project.get('industry', ''),
                            'phases': project.get('phases', []),
                            'headcount': project.get('headcount', ''),
                            'env_langs': project.get('env_langs', ''),
                            'env_tools': project.get('env_tools', ''),
                            'env_dbs': project.get('env_dbs', ''),
                            'env_oss': project.get('env_oss', '')
                        }
                        updated_data['work_content'] = ''
                        for env_type, env_key, years_key in [
                            ("language", "env_langs", "env_langs_years"),
                            ("tool", "env_tools", "env_tools_years"),
                            ("db", "env_dbs", "env_dbs_years"),
                            ("machine", "env_oss", "env_oss_years"),
                        ]:
                            env_val = project.get(env_key, "")
                            years_dict = project.get(years_key, {})
                            env_items = [normalize_skill_name(x) for x in re.split(r"[\n,]", str(env_val)) if normalize_skill_name(x) and normalize_skill_name(x) not in ["''", '""', "[]"]]
                            update_skills_with_env_items(selected_user, env_type, env_items, years_dict)
                        if not update_project(project.get('id'), updated_data):
                            all_success = False
                    if all_success:
                        st.success("すべての案件情報を更新しました！")
                        if "edit_projects_buffer" in st.session_state:
                            del st.session_state["edit_projects_buffer"]
                        st.rerun()
                    else:
                        st.error("案件情報の更新に失敗しました。")
            elif edit_mode == "新しい案件情報の追加":
                st.subheader("新しい案件情報の追加")
                
                if not projects_data.empty:
                    if 'period_start' in projects_data.columns:
                        try:
                            projects_data_sorted = projects_data.copy()
                            projects_data_sorted['period_start_sort'] = pd.to_datetime(projects_data_sorted['period_start'], errors='coerce')
                            projects_data_sorted = projects_data_sorted.sort_values('period_start_sort', ascending=False)
                            projects_list = [row for _, row in projects_data_sorted.iterrows()]
                        except Exception:
                            projects_list = [row for _, row in projects_data.sort_values('id', ascending=False).iterrows()]
                    else:
                        projects_list = [row for _, row in projects_data.sort_values('id', ascending=False).iterrows()]
                    st.write("**現在の案件情報:**")
                    for idx, project in enumerate(projects_list):
                        with st.expander(f"案件 {idx + 1}: {project.get('system_name', 'N/A')}"):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write(f"**期間:** {project.get('period_start', 'N/A')} ～ {project.get('period_end', 'N/A')}")
                                st.write(f"**役割:** {project.get('role', 'N/A')}")
                                st.write(f"**業種:** {project.get('industry', 'N/A')}")
                            with col2:
                                st.write(f"**工程:** {project.get('phases', 'N/A')}")
                                st.write(f"**人数:** {project.get('headcount', 'N/A')}")
                                st.write(f"**作業内容:** {project.get('work_content', 'N/A')}")
                            st.write("**環境情報:**")
                            env_cols = st.columns(4)
                            for i, (env_type, env_label, env_key) in enumerate([
                                ("env_langs", "言語", "language"),
                                ("env_tools", "ツール/FW/Lib", "tool"),
                                ("env_dbs", "DB", "db"),
                                ("env_oss", "OS/マシン", "machine"),
                            ]):
                                with env_cols[i]:
                                    env_display = project.get(env_type, 'N/A')
                                    if isinstance(env_display, dict):
                                        env_display = 'N/A'
                                    st.write(f"**{env_label}:** {env_display}")
                
                st.markdown("---")
                
                with st.form("new_project_form"):
                    col1, col2 = st.columns(2)
                    with col1:
                        def get_default_ym(val):
                            if not val:
                                return ""
                            if isinstance(val, str):
                                try:
                                    if "/" in val:
                                        return datetime.strptime(val, "%Y/%m").strftime("%Y/%m")
                                    elif "-" in val:
                                        return datetime.strptime(val, "%Y-%m").strftime("%Y/%m")
                                    elif len(val) == 6 and val.isdigit():
                                        return f"{val[:4]}/{val[4:]}"
                                except:
                                    return val
                            elif isinstance(val, datetime):
                                return val.strftime("%Y/%m")
                            return str(val)
                        period_start = st.text_input("開始年月 (yyyy/MM)", key="new_period_start")
                        period_end = st.text_input("終了年月 (yyyy/MM)", key="new_period_end")
                        system_name = st.text_input("システム名・作業内容", key="new_system_name")
                        role = st.selectbox("役割", [
                            "CNSL(コンサルタント)", "PMO(プロジェクトマネジメントオフィス)", 
                            "PM(プロジェクトマネージャー)", "PL(プロジェクトリーダー)",
                            "SL(サブリーダー)", "SE(システムエンジニア)", "PG(プログラマ)", "M（メンバー）"
                        ], key="new_role")
                    with col2:
                        industry = st.selectbox("業種", [
                            "農林業", "鉱業", "建設業", "製造業", "情報通信", "運輸業", 
                            "卸売・小売", "金融・保険", "不動産業", "飲食・宿泊所", 
                            "医療・福祉", "教育・学習", "複合サービス事業", "サービス業", "公務", "その他"
                        ], key="new_industry")
                        headcount = st.text_input("人数", key="new_headcount")
                    st.write("**工程:**")
                    phase_cols = st.columns(5)
                    phases = []
                    phase_options = ["環境構築", "要件", "基本", "詳細", "製造", "単体", "結合", "総合", "保守運用", "他"]
                    for i, phase in enumerate(phase_options):
                        with phase_cols[i % 5]:
                            if st.checkbox(phase, key=f"new_phase_{i}"):
                                phases.append(phase)
                    
                    st.write("**環境情報:**")
                    env_cols = st.columns(4)
                    env_inputs = {}
                    env_years_dicts = {}
                    for i, (env_type, env_label, env_key) in enumerate([
                        ("env_langs", "言語", "language"),
                        ("env_tools", "ツール/FW/Lib", "tool"),
                        ("env_dbs", "DB", "db"),
                        ("env_oss", "OS/マシン", "machine"),
                    ]):
                        with env_cols[i]:
                            env_text = st.text_area(env_label, key=f"new_{env_type}", help="複数ある場合は改行またはカンマで区切って入力")
                            env_items = [normalize_skill_name(x) for x in re.split(r"[\n,]", str(env_text)) if normalize_skill_name(x) and normalize_skill_name(x) not in ["''", '""', "[]"]]
                            years_dict_key = f"new_{env_type}_years"
                            if years_dict_key not in st.session_state:
                                st.session_state[years_dict_key] = {x: (0, 0) for x in env_items}
                            years_dict = st.session_state.get(years_dict_key, {})
                            for x in env_items:
                                if x not in years_dict:
                                    # skills_dataから既存年数を取得
                                    years, months = 0, 0
                                    if not skills_data.empty:
                                        skill_row = skills_data[(skills_data['skill_type'] == env_key) & (skills_data['skill_name'] == x)]
                                        if not skill_row.empty:
                                            y, m = float_to_years_months(skill_row.iloc[0]['experience_years'])
                                            years, months = y, m
                                    years_dict[x] = (years, months)
                            for k in list(years_dict.keys()):
                                if k not in env_items:
                                    del years_dict[k]
                            for x in env_items:
                                year_key = f"new_{env_type}_years_{x}_year"
                                month_key = f"new_{env_type}_years_{x}_month"
                                default_year, default_month = years_dict.get(x, (0, 0))
                                y = st.number_input(
                                    f"{env_label}「{x}」経験年数（年）", min_value=0, max_value=50, step=1,
                                    value=st.session_state.get(year_key, default_year), key=year_key
                                )
                                m = st.number_input(
                                    f"{env_label}「{x}」経験年数（月）", min_value=0, max_value=11, step=1,
                                    value=st.session_state.get(month_key, default_month), key=month_key
                                )
                                years_dict[x] = (y, m)
                                st.session_state[years_dict_key] = years_dict
                                st.caption(f"→ {y}年{m}ヶ月")
                            env_inputs[env_type] = env_text
                            env_years_dicts[env_type] = years_dict
                    
                    submitted = st.form_submit_button("案件情報を追加", type="primary")
                    
                    if submitted:
                        def ym_to_date(ym):
                            try:
                                return datetime.strptime(ym, "%Y/%m")
                            except:
                                try:
                                    return datetime.strptime(ym, "%Y-%m")
                                except:
                                    return None

                        def calculate_years_ym(start_ym, end_ym):
                            start = ym_to_date(start_ym)
                            end = ym_to_date(end_ym)
                            if not start or not end:
                                return 0
                            months = (end.year - start.year) * 12 + (end.month - start.month) + 1
                            return round(months / 12, 2)

                        def save_new_project(user_id, project_data):
                            conn = sqlite3.connect(DB_PATH)
                            try:
                                cursor = conn.cursor()
                                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_info'")
                                use_new_schema = cursor.fetchone() is not None
                                if use_new_schema:
                                    cursor.execute("""
                                        INSERT INTO projects (user_info_id, period_start, period_end, system_name,
                                                           role, industry, work_content, phases, headcount,
                                                           env_langs, env_tools, env_dbs, env_oss)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    """, (
                                        user_id, project_data['period_start'], project_data['period_end'],
                                        project_data['system_name'], project_data['role'], project_data['industry'],
                                        '', str(phases) if not isinstance(phases, dict) else '', project_data['headcount'],
                                        project_data['env_langs'], project_data['env_tools'],
                                        project_data['env_dbs'], project_data['env_oss']
                                    ))
                                else:
                                    cursor.execute("""
                                        INSERT INTO projects (basic_info_id, period_start, period_end, system_name,
                                                           role, industry, work_content, phases, headcount,
                                                           env_langs, env_tools, env_dbs, env_oss)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                    """, (
                                        user_id, project_data['period_start'], project_data['period_end'],
                                        project_data['system_name'], project_data['role'], project_data['industry'],
                                        '', str(phases) if not isinstance(phases, dict) else '', project_data['headcount'],
                                        project_data['env_langs'], project_data['env_tools'],
                                        project_data['env_dbs'], project_data['env_oss']
                                    ))
                                conn.commit()
                                return True
                            except Exception as e:
                                conn.rollback()
                                st.error(f"案件情報の保存に失敗しました: {str(e)}")
                                return False
                            finally:
                                conn.close()
                        def update_skills_from_project(user_id, project_data, env_years_dicts):
                            conn = sqlite3.connect(DB_PATH)
                            try:
                                cursor = conn.cursor()
                                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_info'")
                                use_new_schema = cursor.fetchone() is not None
                                if not use_new_schema:
                                    return
                                cursor.execute("SELECT skill_type, skill_name, experience_years FROM skills WHERE user_info_id = ?", (user_id,))
                                all_skills = cursor.fetchall()
                                skill_dict = {}
                                for skill_type, skill_name, exp in all_skills:
                                    norm_name = normalize_skill_name(skill_name)
                                    skill_dict[(skill_type, norm_name)] = exp
                                def process_skill_update(skill_type, env_value, years_dict):
                                    if env_value:
                                        items = [normalize_skill_name(x) for x in re.split(r'[\n,]', str(env_value)) if normalize_skill_name(x) and normalize_skill_name(x) not in ["''", '""', "[]"]]
                                        for item in items:
                                            key = (skill_type, item)
                                            y, m = years_dict.get(item, (0, 0))
                                            years = year_month_to_float(y, m)
                                            if key in skill_dict:
                                                new_years = years  # 上書き方式
                                                cursor.execute("""
                                                    UPDATE skills SET experience_years = ? 
                                                    WHERE user_info_id = ? AND skill_type = ? AND skill_name = ?
                                                """, (new_years, user_id, skill_type, item))
                                            else:
                                                cursor.execute("""
                                                    INSERT INTO skills (user_info_id, skill_type, skill_name, experience_years)
                                                    VALUES (?, ?, ?, ?)
                                                """, (user_id, skill_type, item, years))
                                process_skill_update('language', project_data['env_langs'], env_years_dicts.get('env_langs', {}))
                                process_skill_update('tool', project_data['env_tools'], env_years_dicts.get('env_tools', {}))
                                process_skill_update('db', project_data['env_dbs'], env_years_dicts.get('env_dbs', {}))
                                process_skill_update('machine', project_data['env_oss'], env_years_dicts.get('env_oss', {}))
                                conn.commit()
                                return True
                            except Exception as e:
                                conn.rollback()
                                st.error(f"スキル情報の更新に失敗しました: {str(e)}")
                                return False
                            finally:
                                conn.close()
                        project_data = {
                            'period_start': period_start,
                            'period_end': period_end,
                            'system_name': system_name,
                            'role': role,
                            'industry': industry,
                            'headcount': headcount,
                            'env_langs': env_inputs.get('env_langs', ''),
                            'env_tools': env_inputs.get('env_tools', ''),
                            'env_dbs': env_inputs.get('env_dbs', ''),
                            'env_oss': env_inputs.get('env_oss', '')
                        }
                        project_data['work_content'] = ''
                        if save_new_project(selected_user, project_data):
                            st.success("案件情報を追加しました！")
                            if update_skills_from_project(selected_user, project_data, env_years_dicts):
                                st.success("スキル情報を自動更新しました！")
                            if "edit_projects_buffer" in st.session_state:
                                del st.session_state["edit_projects_buffer"]
                            st.rerun()
                        else:
                            st.error("案件情報の保存に失敗しました。")
            else:
                st.info("このユーザーには案件情報がありません。")
        else:
            st.error("ユーザー情報の取得に失敗しました。")
