import streamlit as st
import openpyxl, shutil, os, datetime, sys
import pandas as pd
import sqlite3
from datetime import datetime as dt
from openpyxl.utils import column_index_from_string, get_column_letter
import base64
import tempfile


# ======== 定数・リスト類（ヘッダー・選択肢など） ========

PHASE_LABELS = [
    "環境構築", "要件", "基本", "詳細", "製造", "単体", "結合", "総合", "保守運用", "他"
]
ROLE_OPTIONS = [
    "CNSL(ｺﾝｻﾙﾀﾝﾄ)", "PMO(ﾌﾟﾛｼﾞｪｸﾄﾏﾈｰｼﾞﾒﾝﾄｵﾌｨｽ)", "PM(ﾌﾟﾛｼﾞｪｸﾄﾏﾈｰｼﾞｬｰ)", "PL(ﾌﾟﾛｼﾞｪｸﾄﾘｰﾀﾞｰ)",
    "SL(ｻﾌﾞﾘｰﾀﾞｰ)", "SE(ｼｽﾃﾑｴﾝｼﾞﾆｱ)", "PG(ﾌﾟﾛｸﾞﾗﾏ)", "M（ﾒﾝﾊﾞｰ）"
]
INDUSTRY_OPTIONS = [
    "農林業", "鉱業", "建設業", "製造業", "情報通信", "運輸業", "卸売・小売", "金融・保険", "不動産業",
    "飲食・宿泊所", "医療・福祉", "教育・学習", "複合サービス事業", "サービス業", "公務", "その他（業務詳細に記入）"
]
ENV_HEADERS = [
    ("言語", "env_langs"),
    ("ﾂｰﾙ\nﾌﾚｰﾑﾜｰｸ\nﾗｲﾌﾞﾗﾘ", "env_tools"),
    ("DB", "env_dbs"),
    ("OS\nﾏｼﾝ", "env_oss"),
]
BASE_ROW_MAP = {
    "period_start": 37,  # D37
    "period_end": 38,    # E38
    "system_and_work": 40,  # J40
    "role": 39,          # T39
    "industry": 39,      # L39
    "headcount": 39,     # Z39
    "phase": 38,         # J38, L38, ... (工程)
    "env_langs": 39,     # AD39
    "env_dbs": 39,       # AJ39
}
COL_MAP = {
    "period_start": "D",
    "period_end": "E",
    "system_and_work": "J",
    "role": "T",
    "industry": "L",
    "headcount": "Z",
    "env_langs": "AD",
    "env_dbs": "AJ",
}
ENV_MAX = 11

b64_SKILLSHEET_TEMPLATE = """
# ここにbase64文字列を貼り付けてください
"""

def get_template_path():
    """
    SkillSheetTemplate.xlsx を Scripts/Template/SkillSheetTemplate.xlsx から優先的に探す。
    それ以外は従来通りのパスもサーチ。
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    # Templateディレクトリを優先
    template_dir = os.path.join(base_dir, "Template")
    template_filename = "SkillSheetTemplate.xlsx"
    path_template = os.path.join(template_dir, template_filename)
    # 旧来のパスも残す
    path1 = os.path.join(base_dir, template_filename)
    path2 = os.path.join(os.getcwd(), template_filename)
    path3 = template_filename
    static_dir = os.path.join(base_dir, "static")
    path4 = os.path.join(static_dir, template_filename)
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    path5 = os.path.join(desktop, template_filename)
    path6 = os.path.join(os.path.dirname(sys.argv[0]), template_filename)
    path7 = "SkillSheetTemplate.xlsx"

    # Template/SkillSheetTemplate.xlsx を最優先
    for p in [path_template, path1, path2, path3, path4, path5, path6, path7]:
        if os.path.exists(p):
            return p

    # base64埋め込みテンプレート
    if b64_SKILLSHEET_TEMPLATE.strip():
        try:
            decoded = base64.b64decode(b64_SKILLSHEET_TEMPLATE.encode() if isinstance(b64_SKILLSHEET_TEMPLATE, str) else b64_SKILLSHEET_TEMPLATE)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.write(decoded)
            tmp.close()
            return tmp.name
        except Exception as e:
            st.error(f"テンプレートの埋め込みデータからの復元に失敗しました: {e}")
            return None
    return None

TEMPLATE_PATH = get_template_path()
OUTPUT_PATH = "SkillSheetOutput.xlsx"
DB_PATH = os.path.join(os.path.dirname(__file__) if '__file__' in globals() else os.getcwd(), "skillsheet_data.db")

def init_database():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_info (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            name_kana TEXT NOT NULL,
            transportation TEXT,
            nearest_station TEXT,
            access_method TEXT,
            access_time TEXT,
            gender TEXT,
            birth_date DATE,
            final_education TEXT,
            graduation_date TEXT,
            self_pr TEXT,
            qualifications TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_info_id INTEGER,
            skill_type TEXT,
            skill_name TEXT,
            experience_years TEXT,
            FOREIGN KEY (user_info_id) REFERENCES user_info (id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT
        )
    ''')
    project_columns = [
        ("user_info_id", "INTEGER"),
        ("period_start", "TEXT"),
        ("period_end", "TEXT"),
        ("system_name", "TEXT"),
        ("role", "TEXT"),
        ("industry", "TEXT"),
        ("work_content", "TEXT"),
        ("phases", "TEXT"),
        ("headcount", "TEXT"),
        ("env_langs", "TEXT"),
        ("env_tools", "TEXT"),
        ("env_dbs", "TEXT"),
        ("env_oss", "TEXT"),
    ]
    cursor.execute("PRAGMA table_info(projects)")
    existing_cols = [row[1] for row in cursor.fetchall()]
    for col, typ in project_columns:
        if col not in existing_cols:
            try:
                cursor.execute(f"ALTER TABLE projects ADD COLUMN {col} {typ}")
            except Exception as e:
                pass
    conn.commit()
    conn.close()

def save_to_database(data):
    conn = sqlite3.connect(DB_PATH, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO user_info (name, name_kana, transportation, nearest_station, 
                                  access_method, access_time, gender, birth_date, 
                                  final_education, graduation_date, self_pr, qualifications)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['name'], data['name_kana'], data['transportation'], data['nearest_station'],
            data['access_method'], data['access_time'], data['gender'], data['birth_date'],
            data['final_education'], data['graduation_date'], data['self_pr'],
            ",".join([q for q in data['qualifications'] if isinstance(q, str) and q.strip()])
        ))
        user_info_id = cursor.lastrowid
        for lang, years in zip(data['languages'], data['language_years']):
            if isinstance(lang, str) and lang.strip():
                cursor.execute('''
                    INSERT INTO skills (user_info_id, skill_type, skill_name, experience_years)
                    VALUES (?, ?, ?, ?)
                ''', (user_info_id, 'language', lang, years))
        for tool, years in zip(data['tools'], data['tool_years']):
            if isinstance(tool, str) and tool.strip():
                cursor.execute('''
                    INSERT INTO skills (user_info_id, skill_type, skill_name, experience_years)
                    VALUES (?, ?, ?, ?)
                ''', (user_info_id, 'tool', tool, years))
        for db, years in zip(data['databases'], data['db_years']):
            if isinstance(db, str) and db.strip():
                cursor.execute('''
                    INSERT INTO skills (user_info_id, skill_type, skill_name, experience_years)
                    VALUES (?, ?, ?, ?)
                ''', (user_info_id, 'db', db, years))
        for machine, years in zip(data['machines'], data['machine_years']):
            if isinstance(machine, str) and machine.strip():
                cursor.execute('''
                    INSERT INTO skills (user_info_id, skill_type, skill_name, experience_years)
                    VALUES (?, ?, ?, ?)
                ''', (user_info_id, 'machine', machine, years))
        for project in data['projects']:
            cursor.execute('''
                INSERT INTO projects (user_info_id, period_start, period_end, system_name,
                                    role, industry, work_content, phases, headcount,
                                    env_langs, env_tools, env_dbs, env_oss)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                user_info_id, project.get('period_start', ""), project.get('period_end', ""),
                project.get('system_and_work', ""), project.get('role', ""), project.get('industry', ""),
                project.get('work_content', ""), str(project.get('phases', "")), project.get('headcount', ""),
                str(project.get('env_langs', "")), str(project.get('env_tools', "")),
                str(project.get('env_dbs', "")), str(project.get('env_oss', ""))
            ))
        conn.commit()
        cursor.execute("PRAGMA wal_checkpoint(FULL)")
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"データベース保存エラー: {str(e)}")
        return False
    finally:
        conn.close()

init_database()

# --- レイアウト調整: サイドバーとセクション分割 ---
st.set_page_config(page_title="スキルシート作成フォーム", layout="wide")
st.markdown(
    """
    <style>
    .block-container {padding-top: 1.5rem;}
    .stTextInput>div>div>input {font-size: 1.1rem;}
    .stTextArea textarea {font-size: 1.1rem;}
    .stSelectbox>div>div>div {font-size: 1.1rem;}
    .stButton>button {font-size: 1.1rem;}
    .stExpanderHeader {font-size: 1.1rem;}
    .sidebar-info-box {
        background: #f5f5f5;
        border-radius: 8px;
        padding: 1em 1em 1em 1em;
        margin-bottom: 1em;
        color: #333;
        font-size: 1.05rem;
        border: 1px solid #e0e0e0;
    }
    .section-box {
        background: #fff;
        border-radius: 10px;
        border: 1.5px solid #e0e0e0;
        padding: 1.2em 1.2em 1.2em 1.2em;
        margin-bottom: 1.5em;
        box-shadow: 0 1px 2px rgba(0,0,0,0.03);
    }
    .section-title {
        font-size: 1.25rem;
        font-weight: bold;
        color: #2c3e50;
        margin-bottom: 0.7em;
        margin-top: 0.2em;
        border: none;
        border-radius: 6px;
        padding: 0.3em 0.8em;
        background: #fff;
        display: inline-block;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    </style>
    """, unsafe_allow_html=True
)

# --- サイドバー: ナビゲーション ---
#with st.sidebar:
   # st.title("📝 スキルシート")
    # メニューや戻るボタンは削除
    # st.markdown("**メニュー**")
    # if st.button("🏠 ホームへ戻る", key="go_home_from_create"):
    #     st.session_state.current_page = "🏠 ホーム"
    #     st.rerun()
    # st.markdown("---")

st.title("スキルシート作成フォーム")
st.markdown("---")

# --- 基本情報セクション ---
with st.container():
    st.markdown('<div class="section-title">基本情報</div>', unsafe_allow_html=True)
    info_cols = st.columns([2, 2, 2])
    name = info_cols[0].text_input("氏名", key="name")
    nameKana = info_cols[1].text_input("カナ", key="nameKana")
    gender = info_cols[2].selectbox("性別", ["男", "女"], key="gender")

    # 交通情報を同じ行に
    st.markdown('<div class="section-title" style="font-size:1.1rem;color:#555;border:none;border-radius:6px;padding:0.2em 0.7em;background:#fff;display:inline-block;">交通情報</div>', unsafe_allow_html=True)
    trans_cols = st.columns([2, 2, 2, 1.2])
    transportation = trans_cols[0].text_input("電車", key="transportation")
    nearest_station = trans_cols[1].text_input("最寄り駅", key="nearest_station")
    access_method = trans_cols[2].selectbox("駅からの交通手段", ["徒歩", "自転車", "バス", "車"], key="access_method")
    access_time = trans_cols[3].text_input("所要時間(分)", max_chars=3, key="access_time", placeholder="分")
    if access_time and not access_time.isdigit():
        st.warning("所要時間は数字のみ入力してください")
        access_time = ""
    st.markdown('</div>', unsafe_allow_html=True)

# --- 生年月日・学歴セクション ---
with st.container():
    st.markdown('<div class="section-title">生年月日・学歴</div>', unsafe_allow_html=True)
    bcols = st.columns([1, 0.15, 0.8, 0.15, 0.8])
    st.markdown("生年月日 (yyyy/MM/dd) ※数字のみ入力してください")
    birth_year = bcols[0].text_input("年", max_chars=4, key="birth_year", placeholder="yyyy")
    bcols[1].markdown("<div style='text-align:center;font-size:24px;'>/</div>", unsafe_allow_html=True)
    birth_month = bcols[2].text_input("月", max_chars=2, key="birth_month", placeholder="MM")
    bcols[3].markdown("<div style='text-align:center;font-size:24px;'>/</div>", unsafe_allow_html=True)
    birth_day = bcols[4].text_input("日", max_chars=2, key="birth_day", placeholder="dd")

    def is_digit_and_len(s, l):
        return s.isdigit() and len(s) == l

    birth_date_str, birth_date, birth_date_valid = "", None, True
    if all([birth_year, birth_month, birth_day]):
        if all([is_digit_and_len(birth_year,4), is_digit_and_len(birth_month,2), is_digit_and_len(birth_day,2)]):
            birth_date_str = f"{birth_year}/{birth_month}/{birth_day}"
            try:
                birth_date = datetime.datetime.strptime(birth_date_str, "%Y/%m/%d").date()
                if birth_date > datetime.date.today():
                    st.warning("生年月日は本日以前の日付を入力してください")
                    birth_date_valid = False
            except:
                st.warning("生年月日は正しい日付を入力してください (例: 1990/01/01)")
                birth_date_valid = False
        else:
            st.warning("生年月日は数字4桁(年)、2桁(月/日)で入力してください")
            birth_date_valid = False
    else:
        birth_date_valid = False

    ecols = st.columns([2, 1, 0.15, 0.8, 0.5])
    st.markdown("最終学歴・卒業年月 (yyyy/MM) ※数字のみ入力してください")
    final_education = ecols[0].text_input("最終学歴", key="final_education")
    graduation_year = ecols[1].text_input("卒業年", max_chars=4, key="graduation_year", placeholder="yyyy")
    ecols[2].markdown("<div style='text-align:center;font-size:24px;'>/</div>", unsafe_allow_html=True)
    graduation_month = ecols[3].text_input("卒業月", max_chars=2, key="graduation_month", placeholder="MM")
    graduation_date_str, graduation_date_valid = "", True
    if graduation_year and graduation_month:
        try:
            m = int(graduation_month)
            if 1 <= m <= 12:
                graduation_date_str = f"{graduation_year}/{graduation_month.zfill(2)}"
                today = datetime.date.today()
                y, mm = int(graduation_year), int(graduation_month)
                if y > today.year or (y == today.year and mm > today.month):
                    st.warning("卒業年月は本日以前の年月を入力してください")
                    graduation_date_valid = False
            else:
                graduation_date_valid = False
        except:
            graduation_date_valid = False
    elif graduation_year or graduation_month:
        graduation_date_valid = False
    st.markdown('</div>', unsafe_allow_html=True)

# --- スキル・資格セクション ---
with st.container():
    st.markdown('<div class="section-title">スキル・資格</div>', unsafe_allow_html=True)
    def dynamic_inputs(label, key_prefix, max_count, year_placeholder=None):
        count_key = f"{key_prefix}_count"
        if count_key not in st.session_state:
            st.session_state[count_key] = 1

        def add():
            st.session_state[count_key] = min(st.session_state[count_key]+1, max_count)
        def remove():
            st.session_state[count_key] = max(st.session_state[count_key]-1, 1)

        st.markdown(f"#### {label}")
        inputs, years = [], []
        # ボタン用のカラム幅を調整
        if year_placeholder:
            cols = st.columns([5, 2, 0.7, 0.7])
        else:
            cols = st.columns([6, 0.7, 0.7])

        # 1つ目
        inputs.append(cols[0].text_input(f"{label}1", key=f"{key_prefix}_1"))
        if year_placeholder:
            years.append(cols[1].text_input(f"経験年数1", key=f"{key_prefix}_year_1", placeholder=year_placeholder))
        with cols[-2]:
            if st.session_state[count_key] < max_count:
                if st.button("＋", key=f"add_{key_prefix}", help=f"{label}入力欄を追加"):
                    add()
        with cols[-1]:
            if st.session_state[count_key] > 1:
                if st.button("－", key=f"remove_{key_prefix}", help=f"{label}入力欄を削除"):
                    remove()
        # 2つ目以降
        for i in range(1, st.session_state[count_key]):
            if year_placeholder:
                c = st.columns([5, 2])
                inputs.append(c[0].text_input(f"{label}{i+1}", key=f"{key_prefix}_{i+1}"))
                years.append(c[1].text_input(f"経験年数{i+1}", key=f"{key_prefix}_year_{i+1}", placeholder=year_placeholder))
            else:
                inputs.append(st.text_input(f"{label}{i+1}", key=f"{key_prefix}_{i+1}"))
        while len(inputs) < max_count:
            inputs.append("")
        while year_placeholder and len(years) < max_count:
            years.append("")
        return (inputs, years) if year_placeholder else (inputs,)

    # 資格
    st.markdown('<div class="section-title" style="font-size:1.1rem;color:#555;border:none;border-radius:6px;padding:0.2em 0.7em;background:#fff;display:inline-block;">資格</div>', unsafe_allow_html=True)
    qualification_inputs, = dynamic_inputs("資格", "qualification", 6)

    # スキル
    st.markdown('<div class="section-title" style="font-size:1.1rem;color:#555;border:none;border-radius:6px;padding:0.2em 0.7em;background:#fff;display:inline-block;">スキル</div>', unsafe_allow_html=True)
    skill_cols = st.columns(2)
    with skill_cols[0]:
        language_inputs, language_years = dynamic_inputs("言語", "language", 10, "例: 3年")
        db_inputs, db_years = dynamic_inputs("DB", "db", 10, "例: 4年")
    with skill_cols[1]:
        tool_inputs, tool_years = dynamic_inputs("ツール/FW/Lib", "tool", 10, "例: 1年")
        machine_inputs, machine_years = dynamic_inputs("マシン/OS", "machine", 10, "例: 5年")

    # 自己PRを一番下に
    st.markdown('<div class="section-title" style="font-size:1.1rem;color:#555;border:none;border-radius:6px;padding:0.2em 0.7em;background:#fff;display:inline-block;">自己PR</div>', unsafe_allow_html=True)
    self_pr = st.text_area("自己PR", height=120, placeholder="自己PRを入力してください", key="self_pr")
    st.markdown('</div>', unsafe_allow_html=True)

# --- 職務経歴（案件）セクション ---
with st.container():
    st.markdown('<div class="section-title" style="border:none;border-radius:6px;padding:0.3em 0.8em;background:#fff;display:inline-block;font-size:1.2rem;font-weight:bold;color:#2c3e50;margin-bottom:0.7em;margin-top:0.2em;">職務経歴（案件）</div>', unsafe_allow_html=True)
    if "project_count" not in st.session_state: st.session_state.project_count = 1
    def add_project(): st.session_state.project_count = min(st.session_state.project_count+1, 10)
    def remove_project(): st.session_state.project_count = max(st.session_state.project_count-1, 1)
    project_btn_cols = st.columns([6, 1, 1])
    if st.session_state.project_count < 10:
        if project_btn_cols[1].button("＋ 案件追加", key="add_project", help="案件入力欄を追加"): add_project()
    if st.session_state.project_count > 1:
        if project_btn_cols[2].button("－ 案件削除", key="remove_project", help="案件入力欄を削除"): remove_project()
    projects = []
    for i in range(st.session_state.project_count):
        with st.expander(f"案件{i+1}", expanded=(i==0)):
            c = st.columns([1,0.2,1,0.2,1])
            start_year = c[0].text_input("開始年", key=f"start_year_{i}", max_chars=4, placeholder="yyyy")
            c[1].markdown("<div style='text-align:center;font-size:20px;'>/</div>", unsafe_allow_html=True)
            start_month = c[2].text_input("開始月", key=f"start_month_{i}", max_chars=2, placeholder="MM")
            c[3].markdown("<div style='text-align:center;font-size:20px;'>〜</div>", unsafe_allow_html=True)
            end_str = c[4].text_input("終了 (yyyy/MM または 現在)", key=f"end_{i}", placeholder="yyyy/MM または 現在")
            system_and_work = st.text_area(
                "システム名/案件名・業務内容",
                key=f"system_and_work_{i}",
                height=100,
                placeholder="例：販売管理システム\n要件定義、設計、開発"
            )
            role_ind_cols = st.columns(2)
            role = role_ind_cols[0].selectbox("役割", ROLE_OPTIONS, key=f"role_{i}")
            industry = role_ind_cols[1].selectbox("業種", INDUSTRY_OPTIONS, key=f"industry_{i}")

            st.markdown("工程")
            phase_cols = st.columns(5)
            phase_values = {label: phase_cols[idx%5].checkbox(label, key=f"phase_{label}_{i}") for idx, label in enumerate(PHASE_LABELS)}
            headcount = st.text_input("人数", key=f"headcount_{i}", placeholder="例: 5名")
            st.markdown("システム環境")

            # 環境入力を横並びでまとめて表示
            env_cols = st.columns(4)
            # 言語
            lang_count_key = f"env_lang_count_{i}"
            if lang_count_key not in st.session_state:
                st.session_state[lang_count_key] = 1
            def add_lang(idx=i):
                st.session_state[lang_count_key] = min(st.session_state[lang_count_key]+1, ENV_MAX)
            def remove_lang(idx=i):
                st.session_state[lang_count_key] = max(st.session_state[lang_count_key]-1, 1)
            env_langs = []
            with env_cols[0]:
                st.markdown("**言語**")
                lang_btn_cols = st.columns([1, 1])
                for j in range(st.session_state[lang_count_key]):
                    env_langs.append(st.text_input(f"言語{j+1}", key=f"env_lang_{i}_{j}", placeholder="例: Python"))
                with lang_btn_cols[0]:
                    if st.session_state[lang_count_key] < ENV_MAX:
                        if st.button("＋", key=f"add_env_lang_{i}", help="言語入力欄を追加"):
                            add_lang()
                with lang_btn_cols[1]:
                    if st.session_state[lang_count_key] > 1:
                        if st.button("－", key=f"remove_env_lang_{i}", help="言語入力欄を削除"):
                            remove_lang()
                while len(env_langs) < ENV_MAX:
                    env_langs.append("")
            # ツール
            tool_count_key = f"env_tool_count_{i}"
            if tool_count_key not in st.session_state:
                st.session_state[tool_count_key] = 1
            def add_tool(idx=i):
                st.session_state[tool_count_key] = min(st.session_state[tool_count_key]+1, ENV_MAX)
            def remove_tool(idx=i):
                st.session_state[tool_count_key] = max(st.session_state[tool_count_key]-1, 1)
            env_tools = []
            with env_cols[1]:
                st.markdown("**ツール/FW/Lib**")
                tool_btn_cols = st.columns([1, 1])
                for j in range(st.session_state[tool_count_key]):
                    env_tools.append(st.text_input(f"ツール/フレームワーク/ライブラリ{j+1}", key=f"env_tool_{i}_{j}", placeholder="例: Django"))
                with tool_btn_cols[0]:
                    if st.session_state[tool_count_key] < ENV_MAX:
                        if st.button("＋", key=f"add_env_tool_{i}", help="ツール/フレームワーク/ライブラリ入力欄を追加"):
                            add_tool()
                with tool_btn_cols[1]:
                    if st.session_state[tool_count_key] > 1:
                        if st.button("－", key=f"remove_env_tool_{i}", help="ツール/フレームワーク/ライブラリ入力欄を削除"):
                            remove_tool()
                while len(env_tools) < ENV_MAX:
                    env_tools.append("")
            # DB
            db_count_key = f"env_db_count_{i}"
            if db_count_key not in st.session_state:
                st.session_state[db_count_key] = 1
            def add_db(idx=i):
                st.session_state[db_count_key] = min(st.session_state[db_count_key]+1, ENV_MAX)
            def remove_db(idx=i):
                st.session_state[db_count_key] = max(st.session_state[db_count_key]-1, 1)
            env_dbs = []
            with env_cols[2]:
                st.markdown("**DB**")
                db_btn_cols = st.columns([1, 1])
                for j in range(st.session_state[db_count_key]):
                    env_dbs.append(st.text_input(f"DB{j+1}", key=f"env_db_{i}_{j}", placeholder="例: MySQL"))
                with db_btn_cols[0]:
                    if st.session_state[db_count_key] < ENV_MAX:
                        if st.button("＋", key=f"add_env_db_{i}", help="DB入力欄を追加"):
                            add_db()
                with db_btn_cols[1]:
                    if st.session_state[db_count_key] > 1:
                        if st.button("－", key=f"remove_env_db_{i}", help="DB入力欄を削除"):
                            remove_db()
                while len(env_dbs) < ENV_MAX:
                    env_dbs.append("")
            # OS/マシン
            os_count_key = f"env_os_count_{i}"
            if os_count_key not in st.session_state:
                st.session_state[os_count_key] = 1
            def add_os(idx=i):
                st.session_state[os_count_key] = min(st.session_state[os_count_key]+1, ENV_MAX)
            def remove_os(idx=i):
                st.session_state[os_count_key] = max(st.session_state[os_count_key]-1, 1)
            env_oss = []
            with env_cols[3]:
                st.markdown("**OS/マシン**")
                os_btn_cols = st.columns([1, 1])
                for j in range(st.session_state[os_count_key]):
                    env_oss.append(st.text_input(f"OS/マシン{j+1}", key=f"env_os_{i}_{j}", placeholder="例: Linux"))
                with os_btn_cols[0]:
                    if st.session_state[os_count_key] < ENV_MAX:
                        if st.button("＋", key=f"add_env_os_{i}", help="OS/マシン入力欄を追加"):
                            add_os()
                with os_btn_cols[1]:
                    if st.session_state[os_count_key] > 1:
                        if st.button("－", key=f"remove_env_os_{i}", help="OS/マシン入力欄を削除"):
                            remove_os()
                while len(env_oss) < ENV_MAX:
                    env_oss.append("")

            start_str = f"{start_year}/{start_month.zfill(2)}" if start_year and start_month else ""
            projects.append(dict(
                period_start=start_str,
                period_end=end_str,
                system_and_work=system_and_work,
                role=role,
                industry=industry,
                phases=phase_values,
                headcount=headcount,
                env_langs=env_langs,
                env_tools=env_tools,
                env_dbs=env_dbs,
                env_oss=env_oss
            ))

def is_filled(val):
    return isinstance(val, str) and val.strip() != ""

name_val = st.session_state.get("name", name)
nameKana_val = st.session_state.get("nameKana", nameKana)
gender_val = st.session_state.get("gender", gender)
birth_date_val = birth_date_str if birth_date_valid else ""

def is_birth_date_valid(val):
    try:
        if isinstance(val, str):
            dt.strptime(val, "%Y/%m/%d")
        elif isinstance(val, datetime.date):
            pass
        else:
            return False
        return True
    except Exception:
        return False

required_fields_filled = (
    is_filled(name_val)
    and is_filled(nameKana_val)
    and gender_val in ["男", "女"]
    and is_birth_date_valid(birth_date_val)
)
if not required_fields_filled:
    st.warning("氏名、カナ、性別、生年月日は必ず入力してください。")

# --- ボタンエリアを横並びで下部に配置 ---
with st.container():
    btn_cols = st.columns([1, 1, 6])
    with btn_cols[0]:
        if st.button("Excelに書き込む"):
            if not required_fields_filled:
                st.warning("氏名、カナ、性別、生年月日は必ず入力してください。")
            elif not TEMPLATE_PATH or not os.path.exists(TEMPLATE_PATH):
                if b64_SKILLSHEET_TEMPLATE.strip():
                    st.info("テンプレートファイルが見つからなかったため、アプリ内蔵のテンプレートを使用します。")
                else:
                    st.error(
                        "テンプレートファイルが見つかりません。\n"
                        "SkillSheetTemplate.xlsx を Scripts/Template フォルダに配置してください。\n"
                        f"現在の検索パス: {TEMPLATE_PATH or '未検出'}"
                    )
            else:
                try:
                    shutil.copy(TEMPLATE_PATH, OUTPUT_PATH)
                    wb = openpyxl.load_workbook(OUTPUT_PATH)
                    ws = wb.active
                    try:
                        ws["AG39"] = ""
                    except Exception:
                        pass
                    ws["E5"], ws["E4"], ws["T5"], ws["AB5"], ws["AJ5"], ws["AM5"], ws["Q5"], ws["AV1"], ws["B8"], ws["AE8"] = (
                        name_val, nameKana_val, transportation, nearest_station, access_method, access_time, gender_val, birth_date_val, final_education, graduation_date_str
                    )
                    def unpack(lst, n): return lst[:n]
                    field_map = [
                        (["B10", "B11", "B12", "V10", "V11", "V12"], None, unpack(qualification_inputs,6), None),
                        (["B15","B16","B17","B18","B19","B20","B21","B22","B23","B24"], ["H15","H16","H17","H18","H19","H20","H21","H22","H23","H24"], unpack(language_inputs,10), unpack(language_years,10)),
                        (["L15","L16","L17","L18","L19","L20","L21","L22","L23","L24"], ["R15","R16","R17","R18","R19","R20","R21","R22","R23","R24"], unpack(tool_inputs,10), unpack(tool_years,10)),
                        (["V15","V16","V17","V18","V19","V20","V21","V22","V23","V24"], ["AB15","AB16","AB17","AB18","AB19","AB20","AB21","AB22","AB23","AB24"], unpack(db_inputs,10), unpack(db_years,10)),
                        (["AF15","AF16","AF17","AF18","AF19","AF20","AF21","AF22","AF23","AF24"], ["AL15","AL16","AL17","AL18","AL19","AL20","AL21","AL22","AL23","AL24"], unpack(machine_inputs,10), unpack(machine_years,10)),
                    ]
                    for value_cells, year_cells, values, years in field_map:
                        for cell, value in zip(value_cells, values):
                            try:
                                ws[cell] = value
                            except AttributeError:
                                pass
                        if years:
                            for cell, year in zip(year_cells, years):
                                try:
                                    ws[cell] = year
                                except AttributeError:
                                    pass
                    ws["B27"] = self_pr

                    def find_cell_by_value(ws, text):
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                            for cell in row:
                                if isinstance(cell.value, str) and cell.value.strip() == text:
                                    return cell
                        return None

                    def write_below(ws, anchor_cell, value, row_offset=2, col_offset=0):
                        if anchor_cell is None:
                            return
                        try:
                            ws.cell(row=anchor_cell.row + row_offset, column=anchor_cell.column + col_offset, value=value)
                        except Exception:
                            pass

                    reordered_projects = []
                    if len(projects) >= 2:
                        first = projects[0]
                        rest = projects[1:]
                        def get_project_keywords(proj):
                            fields = [
                                "system_and_work", "role", "industry", "headcount",
                                "env_langs", "env_tools", "env_dbs", "env_oss"
                            ]
                            keywords = set()
                            for f in fields:
                                v = proj.get(f, "")
                                if isinstance(v, str):
                                    if v.strip():
                                        keywords.add(v.strip())
                                elif isinstance(v, list):
                                    for item in v:
                                        if isinstance(item, str) and item.strip():
                                            keywords.add(item.strip())
                            return keywords
                        first_keywords = get_project_keywords(first)
                        def has_any_keyword(proj, keywords):
                            fields = [
                                "system_and_work", "role", "industry", "headcount",
                                "env_langs", "env_tools", "env_dbs", "env_oss"
                            ]
                            for f in fields:
                                v = proj.get(f, "")
                                if isinstance(v, str):
                                    if v.strip() and v.strip() in keywords:
                                        return True
                                elif isinstance(v, list):
                                    for item in v:
                                        if isinstance(item, str) and item.strip() in keywords:
                                            return True
                            return False
                        matching = [p for p in rest if has_any_keyword(p, first_keywords) and first_keywords]
                        non_matching = [p for p in rest if not has_any_keyword(p, first_keywords) or not first_keywords]
                        if matching:
                            reordered_projects = matching + [first] + non_matching
                        else:
                            reordered_projects = projects
                    else:
                        reordered_projects = projects[:]

                    ENV_ROW_OFFSET = 2
                    ENV_BLOCK_OFFSET = 13
                    for idx, proj in enumerate(reordered_projects):
                        row_offset = idx * ENV_BLOCK_OFFSET
                        ws[f"{COL_MAP['period_start']}{BASE_ROW_MAP['period_start']+row_offset}"] = proj.get("period_start", "")
                        ws[f"{COL_MAP['period_end']}{BASE_ROW_MAP['period_end']+row_offset}"] = proj.get("period_end", "")
                        ws[f"{COL_MAP['system_and_work']}{BASE_ROW_MAP['system_and_work']+row_offset}"] = proj.get("system_and_work", "")
                        ws[f"{COL_MAP['role']}{BASE_ROW_MAP['role']+row_offset}"] = proj.get("role", "")
                        ws[f"{COL_MAP['industry']}{BASE_ROW_MAP['industry']+row_offset}"] = proj.get("industry", "")
                        ws[f"{COL_MAP['headcount']}{BASE_ROW_MAP['headcount']+row_offset}"] = proj.get("headcount", "")
                        env_lang_values = [v for v in proj.get("env_langs", []) if isinstance(v, str) and v.strip()]
                        env_db_values = [v for v in proj.get("env_dbs", []) if isinstance(v, str) and v.strip()]
                        ws[f"{COL_MAP['env_langs']}{BASE_ROW_MAP['env_langs']+row_offset}"] = env_lang_values[0] if env_lang_values else ""
                        ws[f"{COL_MAP['env_dbs']}{BASE_ROW_MAP['env_dbs']+row_offset}"] = env_db_values[0] if env_db_values else ""
                        phase_base_col = column_index_from_string("J")
                        for phase_idx, label in enumerate(PHASE_LABELS):
                            col = phase_base_col + phase_idx * 2
                            cell = f"{get_column_letter(col)}{BASE_ROW_MAP['phase']+row_offset}"
                            mark = "●" if proj.get("phases", {}).get(label) else ""
                            ws[cell] = mark
                        for header_text, env_key in ENV_HEADERS:
                            header = find_cell_by_value(ws, header_text)
                            if header is None:
                                continue
                            base_row = header.row + ENV_ROW_OFFSET + row_offset
                            col = header.column
                            values = proj.get(env_key, [])
                            for i, v in enumerate([val for val in values if isinstance(val, str) and val.strip()][:ENV_MAX]):
                                try:
                                    ws.cell(row=base_row + i, column=col, value=v)
                                except Exception:
                                    pass
                    if reordered_projects:
                        first_proj = reordered_projects[0]
                        ws["D37"] = first_proj.get("period_start", "")
                        ws["E38"] = first_proj.get("period_end", "")
                        ws["J40"] = first_proj.get("system_and_work", "")
                        ws["T39"] = first_proj.get("role", "")
                        ws["L39"] = first_proj.get("industry", "")
                        ws["Z39"] = first_proj.get("headcount", "")
                        first_env_langs = [v for v in first_proj.get("env_langs", []) if isinstance(v, str) and v.strip()]
                        first_env_dbs = [v for v in first_proj.get("env_dbs", []) if isinstance(v, str) and v.strip()]
                        ws["AD39"] = first_env_langs[0] if first_env_langs else ""
                        ws["AJ39"] = first_env_dbs[0] if first_env_dbs else ""
                    proj_ws = wb["Projects"] if "Projects" in wb.sheetnames else wb.create_sheet("Projects")
                    proj_ws.delete_rows(1, proj_ws.max_row)
                    headers = (
                        ["開始(yyyy/MM)","終了(yyyy/MM|現在)","システム名/案件名・業務内容","役割","業種"]
                        + [f"工程:{p}" for p in PHASE_LABELS]
                        + ["人数"]
                        + [f"環境:言語{i+1}" for i in range(ENV_MAX)]
                        + [f"環境:ツール/FW/Lib{i+1}" for i in range(ENV_MAX)]
                        + [f"環境:DB{i+1}" for i in range(ENV_MAX)]
                        + [f"環境:OS/マシン{i+1}" for i in range(ENV_MAX)]
                    )
                    proj_ws.append(headers)
                    for p in reordered_projects:
                        row = [
                            p.get("period_start", ""),
                            p.get("period_end", ""),
                            p.get("system_and_work", ""),
                            p.get("role", ""),
                            p.get("industry", ""),
                            *[1 if p["phases"].get(label) else 0 for label in PHASE_LABELS],
                            p.get("headcount", ""),
                            *p.get("env_langs", [])[:ENV_MAX],
                            *p.get("env_tools", [])[:ENV_MAX],
                            *p.get("env_dbs", [])[:ENV_MAX],
                            *p.get("env_oss", [])[:ENV_MAX],
                        ]
                        proj_ws.append(row)
                    wb.save(OUTPUT_PATH)
                    st.success("Excelに出力しました！")
                    with open(OUTPUT_PATH, "rb") as f:
                        st.download_button("ダウンロード", f, file_name="SkillSheet.xlsx")
                except Exception as e:
                    st.error(f"Excel出力中にエラーが発生しました: {e}")

    with btn_cols[1]:
        if "db_save_btn_disabled" not in st.session_state:
            st.session_state["db_save_btn_disabled"] = False
        save_btn = st.button(
            "データベースに保存",
            disabled=not required_fields_filled or st.session_state["db_save_btn_disabled"],
            key="db_save_btn"
        )
        if save_btn:
            st.session_state["db_save_btn_disabled"] = True
            data = {
                'name': name,
                'name_kana': nameKana,
                'transportation': transportation,
                'nearest_station': nearest_station,
                'access_method': access_method,
                'access_time': access_time,
                'gender': gender,
                'birth_date': birth_date_str if birth_date_valid else None,
                'final_education': final_education,
                'graduation_date': graduation_date_str if graduation_date_valid else None,
                'self_pr': self_pr,
                'qualifications': qualification_inputs,
                'languages': language_inputs,
                'language_years': language_years,
                'tools': tool_inputs,
                'tool_years': tool_years,
                'databases': db_inputs,
                'db_years': db_years,
                'machines': machine_inputs,
                'machine_years': machine_years,
                'projects': projects
            }
            if not required_fields_filled:
                st.error("必須項目をすべて入力してください。")
                st.session_state["db_saved"] = False
                st.session_state["db_save_btn_disabled"] = False
            else:
                result = save_to_database(data)
                if result:
                    st.success("データベースに保存しました！")
                    st.toast("保存が完了しました", icon="✅")
                    st.session_state["db_saved"] = True
                    st.session_state["db_save_btn_disabled"] = False
                else:
                    st.error("データベースの保存に失敗しました。")
                    st.session_state["db_saved"] = False
                    st.session_state["db_save_btn_disabled"] = False
